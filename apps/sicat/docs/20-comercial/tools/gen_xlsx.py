# -*- coding: utf-8 -*-
"""Gera plano-comercial-sicat.xlsx (4 abas) + 3 CSVs espelho a partir de premissas.json.

Uso:  python gen_xlsx.py
Depois recalcular com LibreOffice (openpyxl grava fórmula sem valor):
  & "C:\\Program Files\\LibreOffice\\program\\soffice.exe" --headless --convert-to xlsx --outdir <tmp> ..\\plano-comercial-sicat.xlsx
E validar com:  python verify_xlsx.py <caminho-do-xlsx-recalculado>
"""
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT_DIR = HERE.parent

AMARELO = PatternFill("solid", fgColor="FFF2B2")   # célula EDITÁVEL
CINZA = PatternFill("solid", fgColor="E8E8E8")     # fórmula/derivado
HEADER = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------- premissas
# Layout FIXO da aba Premissas (linha -> chave). verify_xlsx.py importa isto.
PREMISSAS_ROWS = [
    # (linha, rótulo, caminho no json, unidade, editável)
    (2,  "Câmbio USD/BRL",                    ("globais", "cambio_usd_brl", "valor"), "R$/US$", True),
    (3,  "Preço token input (classe Haiku)",  ("ia", "preco_input_usd_mtok", "valor"), "US$/MTok", True),
    (4,  "Preço token output",                ("ia", "preco_output_usd_mtok", "valor"), "US$/MTok", True),
    (5,  "Multiplicador cache read",          ("ia", "cache_read_mult", "valor"), "x input", True),
    (6,  "Multiplicador cache write (info)",  ("ia", "cache_write_mult", "valor"), "x input (amortizado, fora da fórmula)", True),
    (7,  "Cache hit (fração do input)",       ("ia", "cache_hit", "valor"), "0-1", True),
    (8,  "Tokens input por turno",            ("ia", "tokens_input_turno", "valor"), "tokens", True),
    (9,  "Tokens output por turno",           ("ia", "tokens_output_turno", "valor"), "tokens", True),
    (10, "Tokens por arquivo ingerido",       ("ia", "tokens_ingestao_arquivo", "valor"), "tokens", True),
    (11, "Custo hora suporte/CS",             ("custos", "custo_hora_suporte_brl", "valor"), "R$/h", True),
    (12, "Custo fixo operação/mês (sem folha)", ("custos", "custo_fixo_operacao_mes_brl", "valor"), "R$/mês", True),
    (13, "Folha core/mês (linha separada)",   ("custos", "custo_folha_core_mes_brl", "valor"), "R$/mês", True),
    (14, "k_extra (excedente+add-ons)",       ("globais", "k_extra", "valor"), "x MRR base", True),
    (15, "Churn mensal",                      ("globais", "churn_mensal", "valor"), "fração", True),
    (16, "Fiscal: docs/mês por CNPJ",         ("cenarios_mes12", "fiscal_docs_mes_por_cnpj"), "docs", True),
    (17, "Fiscal: preço médio por doc",       ("cenarios_mes12", "fiscal_preco_doc_medio_brl"), "R$/doc", True),
    (18, "Fiscal: mensalidade por CNPJ",      ("addon_fiscal", "mensalidade_cte_mdfe_por_cnpj_brl", "partida"), "R$/CNPJ·mês", True),
]
ROW_IA_TURNO = 20      # fórmula: custo IA por turno (R$)
ROW_IA_INGESTAO = 21   # fórmula: custo IA por ingestão (R$)

# Rateio de infra+storage por cliente (R$/mês) por plano — premissa editável na aba Planos
INFRA_RATEIO = {"essencial": 8, "profissional": 15, "operacional": 30, "enterprise": 60}

# Layout da aba Planos: linha 2..5 = planos, colunas fixas (verify importa)
PLANOS_START_ROW = 2
COL = {
    "nome": "A", "porte": "B", "mensalidade": "C", "franquia": "D", "excedente": "E",
    "contas": "F", "usuarios": "G", "ia_inter": "H", "ia_ingest": "I", "sso": "J",
    "ret_aud": "K", "sup_h": "L", "onb_h": "M", "infra_rateio": "N",
    "rs_por_do": "O", "custo_ia": "P", "custo_sup": "Q", "custo_total": "R", "margem": "S",
}

# Layout da aba Simulador (verify importa): colunas C/D/E = conservador/base/otimista
SIM_COLS = {"conservador": "C", "base": "D", "otimista": "E"}
SIM_ROW_PLANO = {"essencial": 3, "profissional": 4, "operacional": 5, "enterprise": 6}
SIM_ROW_FISCAL_CNPJS = 7
SIM_ROW_MRR_BASE = 9
SIM_ROW_FISCAL = 10
SIM_ROW_MRR_TOTAL = 11
SIM_ROW_CUSTO_VAR = 12
SIM_ROW_CF = 13
SIM_ROW_RESULT = 14
SIM_ROW_RESULT_FOLHA = 15
SIM_ROW_MARGEM_BRUTA = 16
SIM_ROW_BE_SEM_FOLHA = 18
SIM_ROW_BE_COM_FOLHA = 19


def jget(data, path):
    cur = data
    for k in path:
        cur = cur[k]
    return cur


def build(premissas):
    wb = Workbook()

    # ------------------------------------------------ aba 1: Premissas
    ws = wb.active
    ws.title = "Premissas"
    ws["A1"], ws["B1"], ws["C1"] = "Premissa (edite as células AMARELAS)", "Valor", "Unidade / nota"
    for c in ("A1", "B1", "C1"):
        ws[c].font = HEADER
    for row, label, path, unit, editable in PREMISSAS_ROWS:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = jget(premissas, path)
        ws[f"C{row}"] = unit
        if editable:
            ws[f"B{row}"].fill = AMARELO
    ws[f"A{ROW_IA_TURNO}"] = "Custo IA por turno de chat (R$) — fórmula"
    ws[f"B{ROW_IA_TURNO}"] = (
        "=(B8*(1-B7)*B3 + B8*B7*B5*B3 + B9*B4)/1000000*B2"
    )
    ws[f"B{ROW_IA_TURNO}"].fill = CINZA
    ws[f"C{ROW_IA_TURNO}"] = "input não-cacheado + input cache-read + output, em R$"
    ws[f"A{ROW_IA_INGESTAO}"] = "Custo IA por arquivo ingerido (R$) — fórmula"
    ws[f"B{ROW_IA_INGESTAO}"] = "=B10*B3/1000000*B2"
    ws[f"B{ROW_IA_INGESTAO}"].fill = CINZA
    ws[f"C{ROW_IA_INGESTAO}"] = "ingestão = input puro"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 52

    # ------------------------------------------------ aba 2: Linhas de custo
    ws2 = wb.create_sheet("Linhas de custo")
    headers = ["Item", "Tipo", "Mín (R$)", "Máx (R$)", "Confiança", "Fonte"]
    for i, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = HEADER
    linhas = linhas_de_custo(premissas)
    for r, linha in enumerate(linhas, 2):
        for c, val in enumerate(linha, 1):
            ws2.cell(row=r, column=c, value=val)
    for col, w in zip("ABCDEF", (46, 18, 12, 12, 12, 64)):
        ws2.column_dimensions[col].width = w

    # ------------------------------------------------ aba 3: Planos
    ws3 = wb.create_sheet("Planos")
    plan_headers = [
        ("A", "Plano"), ("B", "Porte"), ("C", "Mensalidade (R$) [editável]"),
        ("D", "Franquia DOs/mês"), ("E", "Excedente R$/DO"), ("F", "Contas CETESB"),
        ("G", "Usuários"), ("H", "IA interações/mês"), ("I", "IA ingestões/mês"),
        ("J", "SSO"), ("K", "Retenção auditoria (m)"), ("L", "Suporte h/mês"),
        ("M", "Onboarding h"), ("N", "Infra+storage rateado R$/mês [editável]"),
        ("O", "R$/DO incluso"), ("P", "Custo IA pior caso R$"), ("Q", "Custo suporte R$"),
        ("R", "Custo total/cliente R$"), ("S", "Margem bruta"),
    ]
    for col, h in plan_headers:
        ws3[f"{col}1"] = h
        ws3[f"{col}1"].font = HEADER
        ws3[f"{col}1"].alignment = WRAP
    for i, p in enumerate(premissas["planos"]):
        r = PLANOS_START_ROW + i
        ws3[f"A{r}"] = p["nome"]
        ws3[f"B{r}"] = p["porte"]
        ws3[f"C{r}"] = p["mensalidade_brl"]["partida"]
        ws3[f"C{r}"].fill = AMARELO
        ws3[f"D{r}"] = p["franquia_dos_mes"]
        ws3[f"E{r}"] = p["excedente_por_do_brl"]
        ws3[f"F{r}"] = "ilimitado*" if p["contas_cetesb"] < 0 else p["contas_cetesb"]
        ws3[f"G{r}"] = "ilimitado*" if p["usuarios"] < 0 else p["usuarios"]
        ws3[f"H{r}"] = p["ia_interacoes_mes"]
        ws3[f"I{r}"] = p["ia_ingestoes_mes"]
        ws3[f"J{r}"] = "sim" if p["sso"] else "—"
        ws3[f"K{r}"] = p["retencao_auditoria_meses"]
        ws3[f"L{r}"] = p["suporte_horas_mes"]
        ws3[f"M{r}"] = p["onboarding_horas"]
        ws3[f"N{r}"] = INFRA_RATEIO[p["id"]]
        ws3[f"N{r}"].fill = AMARELO
        ws3[f"O{r}"] = f"=C{r}/D{r}"
        ws3[f"P{r}"] = f"=H{r}*Premissas!B{ROW_IA_TURNO}+I{r}*Premissas!B{ROW_IA_INGESTAO}"
        ws3[f"Q{r}"] = f"=L{r}*Premissas!B11+M{r}*Premissas!B11/12"
        ws3[f"R{r}"] = f"=P{r}+Q{r}+N{r}"
        ws3[f"S{r}"] = f"=(C{r}-R{r})/C{r}"
        ws3[f"S{r}"].number_format = "0.0%"
        for col in "OPQRS":
            ws3[f"{col}{r}"].fill = CINZA
    for i in range(1, 20):
        ws3.column_dimensions[get_column_letter(i)].width = 15
    ws3.column_dimensions["A"].width = 18
    ws3[f"A{PLANOS_START_ROW + 5}"] = (
        "* ilimitado comercial com fair use contratual. Custo IA = pior caso (100% da cota usada); uso real tende a 30-50%."
    )

    # ------------------------------------------------ aba 4: Simulador de receita
    ws4 = wb.create_sheet("Simulador de receita")
    ws4["A1"] = "Simulador (mês 12) — edite os nº de clientes (AMARELO)"
    ws4["A1"].font = HEADER
    ws4["B2"] = "Plano"
    for nome, colw in SIM_COLS.items():
        ws4[f"{colw}2"] = nome.capitalize()
        ws4[f"{colw}2"].font = HEADER
    plano_por_id = {p["id"]: p for p in premissas["planos"]}
    for pid, row in SIM_ROW_PLANO.items():
        ws4[f"B{row}"] = plano_por_id[pid]["nome"]
        for cen, colw in SIM_COLS.items():
            ws4[f"{colw}{row}"] = premissas["cenarios_mes12"][cen][pid]
            ws4[f"{colw}{row}"].fill = AMARELO
    ws4[f"B{SIM_ROW_FISCAL_CNPJS}"] = "CNPJs com add-on fiscal"
    for cen, colw in SIM_COLS.items():
        ws4[f"{colw}{SIM_ROW_FISCAL_CNPJS}"] = premissas["cenarios_mes12"][cen]["fiscal_cnpjs"]
        ws4[f"{colw}{SIM_ROW_FISCAL_CNPJS}"].fill = AMARELO

    labels = {
        SIM_ROW_MRR_BASE: "MRR base (assinaturas)",
        SIM_ROW_FISCAL: "Receita fiscal/mês (add-on)",
        SIM_ROW_MRR_TOTAL: "MRR total (c/ extras k_extra)",
        SIM_ROW_CUSTO_VAR: "Custo variável total",
        SIM_ROW_CF: "Custo fixo (sem folha)",
        SIM_ROW_RESULT: "Resultado (sem folha)",
        SIM_ROW_RESULT_FOLHA: "Resultado (com folha core)",
        SIM_ROW_MARGEM_BRUTA: "Margem bruta do cenário",
        SIM_ROW_BE_SEM_FOLHA: "Break-even MRR (sem folha)",
        SIM_ROW_BE_COM_FOLHA: "Break-even MRR (com folha)",
    }
    for row, label in labels.items():
        ws4[f"B{row}"] = label
        ws4[f"B{row}"].font = HEADER
    first = PLANOS_START_ROW
    last = PLANOS_START_ROW + len(premissas["planos"]) - 1
    for cen, colw in SIM_COLS.items():
        rng_cli = f"{colw}{SIM_ROW_PLANO['essencial']}:{colw}{SIM_ROW_PLANO['enterprise']}"
        ws4[f"{colw}{SIM_ROW_MRR_BASE}"] = (
            f"=SUMPRODUCT({rng_cli},Planos!C{first}:C{last})"
        )
        ws4[f"{colw}{SIM_ROW_FISCAL}"] = (
            f"={colw}{SIM_ROW_FISCAL_CNPJS}*(Premissas!B18+Premissas!B16*Premissas!B17)"
        )
        ws4[f"{colw}{SIM_ROW_MRR_TOTAL}"] = (
            f"={colw}{SIM_ROW_MRR_BASE}*Premissas!B14+{colw}{SIM_ROW_FISCAL}"
        )
        ws4[f"{colw}{SIM_ROW_CUSTO_VAR}"] = (
            f"=SUMPRODUCT({rng_cli},Planos!R{first}:R{last})"
        )
        ws4[f"{colw}{SIM_ROW_CF}"] = "=Premissas!B12"
        ws4[f"{colw}{SIM_ROW_RESULT}"] = (
            f"={colw}{SIM_ROW_MRR_TOTAL}-{colw}{SIM_ROW_CUSTO_VAR}-{colw}{SIM_ROW_CF}"
        )
        ws4[f"{colw}{SIM_ROW_RESULT_FOLHA}"] = (
            f"={colw}{SIM_ROW_RESULT}-Premissas!B13"
        )
        ws4[f"{colw}{SIM_ROW_MARGEM_BRUTA}"] = (
            f"=IF({colw}{SIM_ROW_MRR_TOTAL}=0,0,({colw}{SIM_ROW_MRR_TOTAL}-{colw}{SIM_ROW_CUSTO_VAR})/{colw}{SIM_ROW_MRR_TOTAL})"
        )
        ws4[f"{colw}{SIM_ROW_MARGEM_BRUTA}"].number_format = "0.0%"
        for row in (SIM_ROW_MRR_BASE, SIM_ROW_FISCAL, SIM_ROW_MRR_TOTAL,
                    SIM_ROW_CUSTO_VAR, SIM_ROW_CF, SIM_ROW_RESULT,
                    SIM_ROW_RESULT_FOLHA, SIM_ROW_MARGEM_BRUTA):
            ws4[f"{colw}{row}"].fill = CINZA
            if row != SIM_ROW_MARGEM_BRUTA:
                ws4[f"{colw}{row}"].number_format = "#,##0"
    # break-even usa a margem bruta do cenário BASE
    base_col = SIM_COLS["base"]
    ws4[f"{base_col}{SIM_ROW_BE_SEM_FOLHA}"] = (
        f"=IF({base_col}{SIM_ROW_MARGEM_BRUTA}=0,0,Premissas!B12/{base_col}{SIM_ROW_MARGEM_BRUTA})"
    )
    ws4[f"{base_col}{SIM_ROW_BE_COM_FOLHA}"] = (
        f"=IF({base_col}{SIM_ROW_MARGEM_BRUTA}=0,0,(Premissas!B12+Premissas!B13)/{base_col}{SIM_ROW_MARGEM_BRUTA})"
    )
    for row in (SIM_ROW_BE_SEM_FOLHA, SIM_ROW_BE_COM_FOLHA):
        ws4[f"{base_col}{row}"].fill = CINZA
        ws4[f"{base_col}{row}"].number_format = "#,##0"
    ws4.column_dimensions["B"].width = 34
    for colw in SIM_COLS.values():
        ws4.column_dimensions[colw].width = 15

    return wb


def linhas_de_custo(p):
    """Aba 2 / CSV espelho — item, tipo, mín, máx, confiança, fonte."""
    c = p["custos"]
    f = p["addon_fiscal"]
    rows = []
    for fase in ("lancamento", "crescimento", "escala"):
        d = c["infra_fixa_mes_brl"][fase]
        alcance = "150+ clientes" if d["clientes_ate"] >= 999999 else f"até {d['clientes_ate']} clientes"
        rows.append((f"Infra cloud — fase {fase} ({alcance})",
                     "fixo-mês", d["min"], d["max"], d["confianca"], d["fonte"]))
    rows.append(("Custo fixo de operação (sem folha)", "fixo-mês",
                 c["custo_fixo_operacao_mes_brl"]["valor"], c["custo_fixo_operacao_mes_brl"]["valor"],
                 c["custo_fixo_operacao_mes_brl"]["confianca"], c["custo_fixo_operacao_mes_brl"]["fonte"]))
    rows.append(("Folha core (1 dev + fundador parcial)", "fixo-mês",
                 c["custo_folha_core_mes_brl"]["valor"], c["custo_folha_core_mes_brl"]["valor"],
                 c["custo_folha_core_mes_brl"]["confianca"], c["custo_folha_core_mes_brl"]["fonte"]))
    rows.append(("Suporte/CS (por hora)", "por-cliente-mês",
                 c["custo_hora_suporte_brl"]["valor"], c["custo_hora_suporte_brl"]["valor"],
                 c["custo_hora_suporte_brl"]["confianca"], c["custo_hora_suporte_brl"]["fonte"]))
    ia = p["ia"]
    fx = p["globais"]["cambio_usd_brl"]["valor"]
    t_in = ia["tokens_input_turno"]["valor"]
    t_out = ia["tokens_output_turno"]["valor"]
    p_in = ia["preco_input_usd_mtok"]["valor"]
    p_out = ia["preco_output_usd_mtok"]["valor"]
    rm = ia["cache_read_mult"]["valor"]
    hit = ia["cache_hit"]["valor"]

    def custo_turno(h):
        return round((t_in * (1 - h) * p_in + t_in * h * rm * p_in + t_out * p_out) / 1e6 * fx, 3)

    hit_alto, hit_baixo = min(hit + 0.10, 0.95), max(hit - 0.10, 0.0)
    ingest = round(ia["tokens_ingestao_arquivo"]["valor"] * p_in / 1e6 * fx, 3)
    rows.append((f"IA — custo por turno de chat (faixa = cache hit {hit_alto:.0%} a {hit_baixo:.0%})", "por-interação",
                 custo_turno(hit_alto), custo_turno(hit_baixo), "media",
                 f"tokens {t_in}/{t_out} @ US${p_in}/{p_out} MTok, cache hit central {hit:.0%}, câmbio {fx} — fórmula viva na aba Premissas"))
    rows.append(("IA — custo por arquivo ingerido", "por-interação", ingest, ingest, "media",
                 f"{ia['tokens_ingestao_arquivo']['valor']} tokens de input"))
    rows.append(("Storage de PDFs (por GB·mês)", "por-DO",
                 0.10, 0.25, c["storage_gb_mes_brl"]["confianca"], c["storage_gb_mes_brl"]["fonte"]))
    rows.append(("Banco/fila (por DO)", "por-DO",
                 c["custo_db_por_do_brl"]["valor"], c["custo_db_por_do_brl"]["valor"],
                 c["custo_db_por_do_brl"]["confianca"], c["custo_db_por_do_brl"]["fonte"]))
    rows.append(("Fiscal: custo do parceiro por documento", "referência",
                 f["custo_parceiro_por_doc_brl"]["min"], f["custo_parceiro_por_doc_brl"]["max"],
                 f["custo_parceiro_por_doc_brl"]["confianca"], f["custo_parceiro_por_doc_brl"]["fonte"]))
    rows.append(("Fiscal: mensalidade da plataforma parceira", "referência",
                 f["mensalidade_parceiro_brl"]["min"], f["mensalidade_parceiro_brl"]["max"],
                 f["mensalidade_parceiro_brl"]["confianca"], f["mensalidade_parceiro_brl"]["fonte"]))
    rows.append(("Fiscal: certificado A1 (custo do CLIENTE)", "referência",
                 f["certificado_a1_cliente_brl_ano"]["min"], f["certificado_a1_cliente_brl_ano"]["max"],
                 f["certificado_a1_cliente_brl_ano"]["confianca"], f["certificado_a1_cliente_brl_ano"]["fonte"]))
    rows.append(("Fiscal: dev fase 1 (API parceira, one-off)", "referência",
                 f["dev_fase1_api_parceira"]["custo_brl_min"], f["dev_fase1_api_parceira"]["custo_brl_max"],
                 f["dev_fase1_api_parceira"]["confianca"], f["dev_fase1_api_parceira"]["fonte"]))
    rows.append(("Fiscal: dev emissor próprio (one-off)", "referência",
                 f["dev_emissor_proprio"]["custo_brl_min"], f["dev_emissor_proprio"]["custo_brl_max"],
                 f["dev_emissor_proprio"]["confianca"], f["dev_emissor_proprio"]["fonte"]))
    return rows


def write_csvs(premissas):
    """CSV espelho (UTF-8 BOM) das 3 abas de dados — para diff no git."""
    with open(OUT_DIR / "plano-comercial-sicat.premissas.csv", "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["Premissa", "Valor", "Unidade", "Confiança", "Fonte"])
        for _, label, path, unit, _ in PREMISSAS_ROWS:
            node_path = path[:-1] if path[-1] in ("valor", "partida") else path
            try:
                node = jget(premissas, node_path)
            except (KeyError, TypeError):
                node = {}
            conf = node.get("confianca", "premissa") if isinstance(node, dict) else "premissa"
            fonte = node.get("fonte", "premissa de cenário") if isinstance(node, dict) else "premissa de cenário"
            w.writerow([label, jget(premissas, path), unit, conf, fonte])

    with open(OUT_DIR / "plano-comercial-sicat.linhas-de-custo.csv", "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["Item", "Tipo", "Mín (R$)", "Máx (R$)", "Confiança", "Fonte"])
        w.writerows(linhas_de_custo(premissas))

    with open(OUT_DIR / "plano-comercial-sicat.planos.csv", "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["Plano", "Porte", "Mensalidade partida (R$)", "Mensalidade mín", "Mensalidade máx",
                    "Franquia DOs/mês", "Excedente R$/DO", "Contas CETESB", "Usuários",
                    "IA interações/mês", "IA ingestões/mês", "SSO", "Retenção auditoria (m)"])
        for p in premissas["planos"]:
            w.writerow([
                p["nome"], p["porte"], p["mensalidade_brl"]["partida"],
                p["mensalidade_brl"]["min"], p["mensalidade_brl"]["max"],
                p["franquia_dos_mes"], p["excedente_por_do_brl"],
                "ilimitado" if p["contas_cetesb"] < 0 else p["contas_cetesb"],
                "ilimitado" if p["usuarios"] < 0 else p["usuarios"],
                p["ia_interacoes_mes"], p["ia_ingestoes_mes"],
                "sim" if p["sso"] else "não", p["retencao_auditoria_meses"],
            ])


def main():
    premissas = json.loads((HERE / "premissas.json").read_text(encoding="utf-8"))
    wb = build(premissas)
    out = OUT_DIR / "plano-comercial-sicat.xlsx"
    wb.save(out)
    write_csvs(premissas)
    print(f"OK: {out}")
    print("OK: 3 CSVs espelho gerados")
    print("Agora recalcule com soffice e rode verify_xlsx.py (ver README.md)")


if __name__ == "__main__":
    main()
