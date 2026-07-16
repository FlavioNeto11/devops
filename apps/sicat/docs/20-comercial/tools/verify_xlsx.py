# -*- coding: utf-8 -*-
"""Golden check: recomputa os números do simulador em Python puro (a partir de
premissas.json) e compara com as células do xlsx RECALCULADO pelo LibreOffice.

Uso:  python verify_xlsx.py <caminho-do-xlsx-recalculado>
Falha (exit 1) se: divergência > 0.5%, erro de fórmula (#REF!/#NAME?/#DIV) ou célula vazia.
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

import gen_xlsx as G

HERE = Path(__file__).resolve().parent
TOL = 0.005  # 0,5%

ERROS_FORMULA = ("#REF!", "#NAME?", "#DIV/0!", "#VALUE!", "#N/A", "#NULL!", "#NUM!")


def esperado(p):
    """Reimplementa as fórmulas da planilha em Python puro."""
    fx = p["globais"]["cambio_usd_brl"]["valor"]
    ia = p["ia"]
    t_in = ia["tokens_input_turno"]["valor"]
    t_out = ia["tokens_output_turno"]["valor"]
    hit = ia["cache_hit"]["valor"]
    p_in = ia["preco_input_usd_mtok"]["valor"]
    p_out = ia["preco_output_usd_mtok"]["valor"]
    read_mult = ia["cache_read_mult"]["valor"]
    ia_turno = (t_in * (1 - hit) * p_in + t_in * hit * read_mult * p_in + t_out * p_out) / 1e6 * fx
    ia_ingest = ia["tokens_ingestao_arquivo"]["valor"] * p_in / 1e6 * fx

    custo_hora = p["custos"]["custo_hora_suporte_brl"]["valor"]
    custo_cliente = {}
    mensalidade = {}
    for plano in p["planos"]:
        pid = plano["id"]
        mensalidade[pid] = plano["mensalidade_brl"]["partida"]
        c_ia = plano["ia_interacoes_mes"] * ia_turno + plano["ia_ingestoes_mes"] * ia_ingest
        c_sup = plano["suporte_horas_mes"] * custo_hora + plano["onboarding_horas"] * custo_hora / 12
        custo_cliente[pid] = c_ia + c_sup + G.INFRA_RATEIO[pid]

    k_extra = p["globais"]["k_extra"]["valor"]
    cf = p["custos"]["custo_fixo_operacao_mes_brl"]["valor"]
    folha = p["custos"]["custo_folha_core_mes_brl"]["valor"]
    fiscal_mens = p["addon_fiscal"]["mensalidade_cte_mdfe_por_cnpj_brl"]["partida"]
    fdocs = p["cenarios_mes12"]["fiscal_docs_mes_por_cnpj"]
    fdoc_preco = p["cenarios_mes12"]["fiscal_preco_doc_medio_brl"]

    out = {"ia_turno": ia_turno, "ia_ingestao": ia_ingest, "cenarios": {}}
    for cen in ("conservador", "base", "otimista"):
        cli = p["cenarios_mes12"][cen]
        mrr_base = sum(cli[pid] * mensalidade[pid] for pid in mensalidade)
        fiscal = cli["fiscal_cnpjs"] * (fiscal_mens + fdocs * fdoc_preco)
        mrr_total = mrr_base * k_extra + fiscal
        custo_var = sum(cli[pid] * custo_cliente[pid] for pid in custo_cliente)
        margem = 0 if mrr_total == 0 else (mrr_total - custo_var) / mrr_total
        out["cenarios"][cen] = {
            "mrr_base": mrr_base, "fiscal": fiscal, "mrr_total": mrr_total,
            "custo_var": custo_var, "resultado": mrr_total - custo_var - cf,
            "resultado_folha": mrr_total - custo_var - cf - folha, "margem": margem,
        }
    m_base = out["cenarios"]["base"]["margem"]
    out["be_sem_folha"] = 0 if m_base == 0 else cf / m_base
    out["be_com_folha"] = 0 if m_base == 0 else (cf + folha) / m_base
    out["custo_cliente"] = custo_cliente
    out["mensalidade"] = mensalidade
    return out


def cmp(nome, got, want, falhas):
    if got is None:
        falhas.append(f"{nome}: célula vazia (recalc não rodou?)")
        return
    if isinstance(got, str):
        falhas.append(f"{nome}: valor não-numérico {got!r}")
        return
    ref = max(abs(want), 1e-9)
    if abs(got - want) / ref > TOL:
        falhas.append(f"{nome}: planilha={got:.4f} esperado={want:.4f}")


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(2)
    xlsx = Path(sys.argv[1])
    premissas = json.loads((HERE / "premissas.json").read_text(encoding="utf-8"))
    want = esperado(premissas)
    wb = load_workbook(xlsx, data_only=True)

    falhas = []

    # 1) varredura de erros de fórmula em todas as abas
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in ERROS_FORMULA:
                    falhas.append(f"{ws.title}!{cell.coordinate}: erro de fórmula {cell.value}")

    # 2) golden check
    prem = wb["Premissas"]
    cmp("Premissas!custo_ia_turno", prem[f"B{G.ROW_IA_TURNO}"].value, want["ia_turno"], falhas)
    cmp("Premissas!custo_ia_ingestao", prem[f"B{G.ROW_IA_INGESTAO}"].value, want["ia_ingestao"], falhas)

    planos_ws = wb["Planos"]
    for i, plano in enumerate(premissas["planos"]):
        r = G.PLANOS_START_ROW + i
        pid = plano["id"]
        m = want["mensalidade"][pid]
        cmp(f"Planos!{pid}.custo_total", planos_ws[f"R{r}"].value, want["custo_cliente"][pid], falhas)
        cmp(f"Planos!{pid}.margem", planos_ws[f"S{r}"].value, (m - want["custo_cliente"][pid]) / m, falhas)

    sim = wb["Simulador de receita"]
    for cen, colw in G.SIM_COLS.items():
        w = want["cenarios"][cen]
        cmp(f"{cen}.mrr_base", sim[f"{colw}{G.SIM_ROW_MRR_BASE}"].value, w["mrr_base"], falhas)
        cmp(f"{cen}.fiscal", sim[f"{colw}{G.SIM_ROW_FISCAL}"].value, w["fiscal"], falhas)
        cmp(f"{cen}.mrr_total", sim[f"{colw}{G.SIM_ROW_MRR_TOTAL}"].value, w["mrr_total"], falhas)
        cmp(f"{cen}.custo_var", sim[f"{colw}{G.SIM_ROW_CUSTO_VAR}"].value, w["custo_var"], falhas)
        cmp(f"{cen}.resultado", sim[f"{colw}{G.SIM_ROW_RESULT}"].value, w["resultado"], falhas)
        cmp(f"{cen}.resultado_folha", sim[f"{colw}{G.SIM_ROW_RESULT_FOLHA}"].value, w["resultado_folha"], falhas)
        cmp(f"{cen}.margem", sim[f"{colw}{G.SIM_ROW_MARGEM_BRUTA}"].value, w["margem"], falhas)
    base_col = G.SIM_COLS["base"]
    cmp("be_sem_folha", sim[f"{base_col}{G.SIM_ROW_BE_SEM_FOLHA}"].value, want["be_sem_folha"], falhas)
    cmp("be_com_folha", sim[f"{base_col}{G.SIM_ROW_BE_COM_FOLHA}"].value, want["be_com_folha"], falhas)

    if falhas:
        print("GOLDEN CHECK FALHOU:")
        for f in falhas:
            print(f"  - {f}")
        sys.exit(1)

    print("GOLDEN CHECK OK — números da planilha batem com premissas.json")
    print(f"  custo IA/turno       R$ {want['ia_turno']:.4f}")
    print(f"  custo IA/ingestão    R$ {want['ia_ingestao']:.4f}")
    for cen in ("conservador", "base", "otimista"):
        w = want["cenarios"][cen]
        print(f"  {cen:<12} MRR total R$ {w['mrr_total']:>10,.0f} | resultado s/folha R$ {w['resultado']:>10,.0f} | margem {w['margem']:.1%}")
    print(f"  break-even MRR sem folha R$ {want['be_sem_folha']:,.0f} | com folha R$ {want['be_com_folha']:,.0f}")


if __name__ == "__main__":
    main()
